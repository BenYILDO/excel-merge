import os
from copy import copy
from io import BytesIO

import pandas as pd
import streamlit as st
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    st.error(
        "'openpyxl' kütüphanesi yüklü değil. Lütfen `pip install openpyxl` komutuyla kurun."
    )
    st.stop()

st.title("Excel Dosyalarını Birleştir")

uploaded_files = st.file_uploader(
    "Excel dosyalarını yükleyin (.xls veya .xlsx)",
    type=["xls", "xlsx"],
    accept_multiple_files=True,
)

has_header = st.checkbox("Dosyalarda başlık satırı var", value=True)

if uploaded_files and st.button("Birleştir"):
    first_file = uploaded_files[0]
    ext = os.path.splitext(first_file.name)[1].lower()
    keep_styles = ext == ".xlsx"

    if keep_styles:
        merged_wb = load_workbook(BytesIO(first_file.getvalue()))
        merged_ws = merged_wb.active
    else:
        df = pd.read_excel(
            BytesIO(first_file.getvalue()),
            header=0 if has_header else None,
            engine="xlrd",
        )
        merged_wb = Workbook()
        merged_ws = merged_wb.active
        if has_header:
            merged_ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            merged_ws.append(list(row))

    for file in uploaded_files[1:]:
        ext = os.path.splitext(file.name)[1].lower()
        if ext == ".xlsx":
            wb = load_workbook(BytesIO(file.getvalue()))
            ws = wb.active
            start = 2 if has_header else 1
            for row in ws.iter_rows(min_row=start, values_only=False):
                target_row = merged_ws.max_row + 1
                for cell in row:
                    new_cell = merged_ws.cell(row=target_row, column=cell.col_idx, value=cell.value)
                    if keep_styles and cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
        else:
            df = pd.read_excel(
                BytesIO(file.getvalue()),
                header=0 if has_header else None,
                engine="xlrd",
            )
            for row in df.itertuples(index=False):
                merged_ws.append(list(row))

    output = BytesIO()
    merged_wb.save(output)
    output.seek(0)

    st.download_button(
        label="Birleştirilmiş Dosyayı İndir",
        data=output,
        file_name="merged.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
